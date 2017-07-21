#!/usr/bin/env python3

__author__ = 	"Marat Strelets"
__copyright__ = "Copyright 2017"
__version__ = 	"0.2"
__email__ = 	"marat.strelets@gmail.com"
__status__ = 	"Beta"

# Imports
import sys
import openpyxl
import datetime
import time
import os 
import jinja2
import re
import selenium.webdriver
import selenium.webdriver.chrome.options
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from email import encoders
import argparse
import logging
from selenium.webdriver.remote.remote_connection import LOGGER
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

# Environment
DIR_PATH = os.path.dirname(os.path.realpath(__file__))

# CONSTS
RESULTS_FILE = "kijiji-scraper-results.xlsx"
LOG_FILE = "kijiji-scraper.log"

# Locators
AD_IN_LIST = '//*[@class = "title"]/a'
BREADCRUMB = '//div[@id = "Breadcrumb"]//a'
RESULTS_HEADER = '//div[contains(@class, "showing")]'
AD_FIELDS = {
	"Title" : "//h1",
	"Description" : "//span[@itemprop = 'description']",
	"Visits" : "//span[@class = 'ad-visits']"
}
AD_OFFER_PRICE = "//span[@itemprop='price']"
AD_CSS_PRICE = "//span[starts-with(@class, 'currentPrice-')]"
AD_CSS_VISITS = "//div[starts-with(@class, 'visitCounter-')]/span"
AD_CSS_DESCRIPTION = "//div[starts-with(@class, 'descriptionContainer-')]"
AD_ATTRIBUTES = "//table[contains(@class, 'ad-attributes')]/tbody/tr"

# Global Variables
args = None
driver = None
wb = None
ws = None
columns = []
excel_row_index = 1

# Functions
def run(url):
	global wb, ws, driver

	# Init start time
	start = datetime.datetime.now().replace(microsecond=0)

	# Init Log
	log_file = open(LOG_FILE, 'w+')
	log_file.close()

	# Init Driver
	LOGGER.setLevel(logging.DEBUG)
	driver = init_driver()

	if driver is None:
		sys.exit()
		
	# Init Excel
	wb = openpyxl.Workbook()
	ws = wb.active

	# TODO: Check if max page exists 

	# Init Scraping
	original_url = url
	proceed = True
	try:
		driver.get(original_url)
		title = driver.title
	except Exception as e:
		log("ERROR Failed openning " + original_url + " - " + str(e).split(os.linesep, 1)[0])
		close_driver()
		proceed = False

	if proceed is True:

		# Get all URLs
		ads_urls = []
		for page_index in range(1, args.pages + 1):

			if driver is None:
				driver = init_driver()
				if driver is None:
					log("WARN  Failed to initialize driver, page " + page_index + " will be skipped")	
					continue

			index = original_url.rfind('/')
			url = original_url[:index] + '/page-' + str(page_index) + original_url[index:]

			page_urls = get_page_ads(url,page_index)

			if len(page_urls) == 0:
				continue

			for ad_url in page_urls:
				if ad_url not in ads_urls:
					ads_urls.append(ad_url)

			if (page_index) % 10 == 0 or page_index == args.pages:
				close_driver()

		# Parse ads
		total = len(ads_urls)
		log("INFO  * Total of " + str(total) + " ads will be scraped")

		for index, ad_url in enumerate(ads_urls):

			if driver is None:
				driver = init_driver()

			log("INFO  Parsing " + str(index + 1) + "/" + str(total) + ": " + ad_url)
			try:
				parse_ad(ad_url)
			except selenium.common.exceptions.TimeoutException as e:
				log("ERROR Failed to load page " + str(page_index) + " before time out elapsed, closing driver...")
				close_driver()
			except Exception as e:
				log("ERROR Failed to parse " + ad_url + " - (" + str(e).split(os.linesep, 1)[0] + ")")

			if (index + 1) % 20 == 0 or index + 1 == total:
				close_driver()

		# Save Results to Excel
		try:
			if os.path.isfile(RESULTS_FILE):
				os.remove(RESULTS_FILE)
			wb.save(RESULTS_FILE)
			log("INFO  Results saved as " + RESULTS_FILE + " in " + DIR_PATH)
		except Exception as e:
			log("ERROR Failed to save results " + str(e))
		

	# Calculate Duration
	end = datetime.datetime.now().replace(microsecond=0)

	# Send eMail
	if args.mail is True and proceed is True:

		if can_send_email():
			from_name = "Kijiji Scrapper"
			if os.path.isfile(RESULTS_FILE):
				message = title + os.linesep
				message += original_url + os.linesep
				message += "Duration - " + str(end - start) + os.linesep
				send_mail(from_name, args.recipients, "Kijiji Scraping Results", message, [RESULTS_FILE, LOG_FILE], args.smtp_server, args.smtp_server_port, args.smtp_server_username, args.smtp_server_password, True)
			else:
				message = "Operation Failed." + os.linesep
				message += "Logs Attached" + os.linesep
				message += "Duration - " + str(end - start)
				send_mail(from_name, args.recipients, "Kijiji Scraping Results - Failed", "Logs attached", [LOG_FILE], args.smtp_server, args.smtp_server_port, args.smtp_server_username, args.smtp_server_password, True)
			log("INFO  Sent results by mail to: " + ','.join(args.recipients))
		else:
			log("ERROR Can not send eMail")

	log("INFO  Script finsihed in " + str(end - start))

	# Cleanup
	if (args.mail):
		if os.path.exists(RESULTS_FILE):
			os.remove(RESULTS_FILE)
		if os.path.exists(LOG_FILE):
			os.remove(LOG_FILE)

	print("Bye.")

def init_driver():
	global driver

	log("INFO  Initializing driver...")
	
	driver_executable = None
	if sys.platform == 'win32':
		driver_executable = "chromedriver.exe"
	else:
		driver_executable = "chromedriver"

	chrome_options = webdriver.ChromeOptions()

	try:
		if args.server is not None: # Selenium Server	

			if args.no_optimize is True:
				driver = webdriver.Remote(
					command_executor=args.server, 
					desired_capabilities=chrome_options.to_capabilities())
			else:
				
				prefs = {"profile.managed_default_content_settings.images":2}
				chrome_options.add_experimental_option("prefs",prefs)
				driver = webdriver.Remote(
					command_executor=args.server, 
					desired_capabilities=chrome_options.to_capabilities())

		else: # Standalone Driver		
			if args.no_optimize is True:
				driver = selenium.webdriver.Chrome(DIR_PATH + "/" + driver_executable)
			else:
				prefs = {"profile.managed_default_content_settings.images":2}
				chrome_options.add_experimental_option("prefs",prefs)

				if args.headless is True:
					chrome_options.add_argument("--headless")
					chrome_options.add_argument("--disable-infobars")
					chrome_options.add_argument("--enable-automation")

				driver = selenium.webdriver.Chrome(DIR_PATH + "/" + driver_executable, chrome_options=chrome_options)
	except Exception as e:
			log("ERROR Failed to initialize driver: " + str(e).split(os.linesep, 1)[0])

	if driver is not None:
		driver.set_page_load_timeout(args.timeout)
		driver.maximize_window()

	return driver

def close_driver():
	global driver

	if driver is not None:
		log("INFO  Closing driver...")
		try:
			driver.close()
			driver = None
		except:
			log("ERROR Failed to close driver")

def prepare_url(url, page_number):
	template = jinja2.Template(url)
	result = template.render(page="/page-"+str(page_number))
	log("INFO  Page " + str(page_number) + " URL: " + result)
	return result

def extract_ad_attribute(element):

	try:
		key = element.find_element_by_tag_name('th').text.replace(os.linesep, ' ').strip()
		value = element.find_element_by_tag_name('td').text.replace(os.linesep, ' ').strip()
	except:
		return None

	if value == "":
		return None

	return {
		"key" : key,
		"value" : value
	}

def get_page_ads(page_url, page_index):

	global driver

	log("INFO  Extracting Ads from Page " + str(page_index) + ": " + page_url)

	urls = []
	
	# Get Data
	try:

		# Load Page
		# FIXME: Kijiji redirects to last page if page > max, (e.g. 30 instead of 100 - wasting time!)
		driver.get(page_url)

		# Get Header
		header = driver.find_element_by_xpath(RESULTS_HEADER).text.strip()
		log("INFO  " + header)

		# Get Ads URLs
		for ad in driver.find_elements_by_xpath(AD_IN_LIST):
			urls.append(ad.get_attribute('href'))

	except selenium.common.exceptions.TimeoutException as e:
		log("ERROR Failed to load page " + str(page_index) + " before time out elapsed, closing driver...")
		close_driver()
	except Exception as e:
		log("ERROR Failed to extract ads on page " + str(page_index) + ": " + str(e).split(os.linesep, 1)[0])

	return urls

def parse_ad(url):
	global driver

	# Navigate to URL
	driver.get(url)

	ad_info = {}
	
	ad_info["URL"] = url

	# Parse Breadcrumb
	breadcrumb = driver.find_elements_by_xpath(BREADCRUMB)
	breadcrumb_index = 1
	for element in breadcrumb:
		key = 'Breadcrumb ' + str(breadcrumb_index)
		breadcrumb_index += 1
		ad_info[key] = element.text.strip()
		log("INFO  ~ " + key + ": " + ad_info[key])

	# Extract AD static fields
	for key in AD_FIELDS:
		try:			
			element = driver.find_element_by_xpath(AD_FIELDS[key])
			ad_info[key] = element.text.replace(os.linesep, ' ').strip()

			if len(ad_info[key]) < 64:
				log("INFO  - " + key + ": " + ad_info[key])
			else:
				log("INFO  - " + key + ": " + ad_info[key][:61] + "...")
		except Exception as e:
			log("WARN  ! Failed to retrieve " + key, True) 

	# Extract AD dynamic fields
	try:
		attributes = driver.find_elements_by_xpath(AD_ATTRIBUTES)
		for attribute in attributes:
			pair = extract_ad_attribute(attribute)
			if pair is not None:
				key = pair["key"]
				value = pair["value"]

				ad_info[key] = value
				if len(value) < 64:
					log("INFO  - " + key + ": " + value)
				else:
					log("INFO  - " + key + ": " + value[:61] + "...")
	except Exception as e:
		log("ERROR Failed to get AD attributes " + str(e))

	# Special care for offers Price (2)
	if not "Price" in ad_info:
		try:
			element = attributes = driver.find_element_by_xpath(AD_OFFER_PRICE)
			ad_info["Price"] = element.text.strip()
			log("INFO  - Price: " + ad_info["Price"])
		except:
			log("WARN  ! Failed to retrieve offer Price (2)", True)

	# Special care for Offer Price (3)
	if not "Price" in ad_info:
		try:
			element = attributes = driver.find_element_by_xpath(AD_CSS_PRICE)
			ad_info["Price"] = element.text.strip()
			log("INFO  - Price: " + ad_info["Price"])
		except:
			log("WARN  ! Failed to retrieve offer Price (3)")

	# Special care for Visits
	if not "Visits" in ad_info:
		try:
			element = attributes = driver.find_element_by_xpath(AD_CSS_VISITS)
			ad_info["Visits"] = element.text.replace('visits', '').strip()
			log("INFO  - Visits: " + ad_info["Visits"])
		except:
			log("WARN  ! Failed to retrieve offer Visits (2)")

	# Special care for Description
	if not "Description" in ad_info:
		try:
			element = attributes = driver.find_element_by_xpath(AD_CSS_DESCRIPTION)
			ad_info["Description"] = element.text.replace(os.linesep, ' ').strip()
			if len(ad_info["Description"]) < 64:
					log("INFO  - Description: " + ad_info["Description"])
			else:
				log("INFO  - Description: " + ad_info["Description"][:61] + "...")
		except:
			log("WARN  ! Failed to retrieve offer Description (2)")


	# Numberilize Visits
	if "Visits" in ad_info and len(ad_info["Visits"]) > 0:
		original_value = ad_info["Visits"]
		new_value = re.sub("[^0-9]", "", original_value)
		if new_value != original_value:
			ad_info["Visits"] = new_value
			log("INFO  + Vists = '" + original_value + "' >> Visits = '" + new_value + "'")

	try:
		save_ad_info(ad_info)
	except Exception as e:
		log("ERROR Failed to save AD info to Excel " + str(e))

def save_ad_info(ad_info):
	global excel_row_index, ws, columns
	excel_row_index += 1

	for key in ad_info:
		# Append new columns
		if not key in columns:
			# Save new Column
			columns.append(key)
			column_index = columns.index(key) + 1

			# Save Header
			ws.cell(row=1, column=column_index).value = key

		# Find out the index of the column
		column_index = columns.index(key) + 1

		# Save in WorkSheet
		ws.cell(row=excel_row_index, column=column_index).value = ad_info[key]

def log(text, skip_print = False):
	if not skip_print:
		print(text)

	date_time = datetime.datetime.now()
	log_file = open(LOG_FILE, 'a')
	date_time = str(date_time)
	text = date_time + '\t' + text + '\r\n'
	log_file.write(text)
	log_file.close()

def send_mail( send_from, send_to, subject, text, files=[], server="localhost", port=587, username='', password='', isTls=True):
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime = True)
    msg['Subject'] = subject

    msg.attach( MIMEText(text) )

    for f in files:
        part = MIMEBase('application', "octet-stream")
        part.set_payload( open(f,"rb").read() )
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="{0}"'.format(os.path.basename(f)))
        msg.attach(part)

    smtp = smtplib.SMTP(server, port)
    if isTls: smtp.starttls()
    smtp.login(username,password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()

def verify_driver():
	message = "ERROR:  ChromeDriver is missing, download from https://sites.google.com/a/chromium.org/chromedriver/downloads"
	if sys.platform == 'win32':
		if not os.path.exists(DIR_PATH + '/chromedriver.exe'):
			log(message)
			sys.exit()
	else:
		if not os.path.exists(DIR_PATH + '/chromedriver'):
			log(message)
			sys.exit()

def can_send_email():
	if args.recipients is None or len(args.recipients) == 0:
		log("WARN  No recipients specified")
		return False

	if args.smtp_server is None:
		log("WARN  No SMTP server provided")
		return  False

	if args.smtp_server_port is None:
		log("WARN  No SMTP server port provided")
		return  False

	if args.smtp_server_username is None:
		log("WARN  No SMTP server username provided")
		return False

	if args.smtp_server_password is None:
		log("WARN  No SMTP server password provided")
		return False

	return True

def setup_args():
	global args
	parser = argparse.ArgumentParser()
	parser.add_argument("source", type=str, help="Source URL to scrape")
	parser.add_argument("-p", "--pages", type=int, default=1, help="Total pages to scrape", metavar="")
	parser.add_argument("-t", "--timeout", type=int, default=60, help="Max timeout (seconds) for Chrome Driver operation", metavar="")
	parser.add_argument("-o", "--no-optimize", dest='no_optimize', action='store_true', help="Disable optimize (load images)")
	parser.add_argument("-e", "--headless", dest='headless', action='store_true', help="Headless mode (no UI)")
	parser.add_argument("-s", "--server", dest='server', type=str, help="Selenium Server URL", metavar="")
	parser.add_argument("-m", "--mail", dest='mail', action='store_true', help="Send email when done")
	parser.add_argument("--smtp-server", dest='smtp_server', type=str, help="SMTP server", metavar="")
	parser.add_argument("--smtp-server-port", dest='smtp_server_port', type=int, help="SMTP server port", metavar="")
	parser.add_argument("--smtp-server-username", dest='smtp_server_username', type=str, help="SMTP server username", metavar="")
	parser.add_argument("--smtp-server-password", dest='smtp_server_password', type=str, help="SMTP server password", metavar="")
	parser.add_argument("-r", "--recipients", nargs='*', default='', type=str, help="Email recipients list", metavar="")
	args = parser.parse_args()

if __name__ == "__main__":
	setup_args()

	log("INFO  Kijiji Scrapper v" + __version__)
	log("INFO  Running on " + sys.platform)

	verify_driver()

	log("INFO  Starting from " + args.source)

	if args.pages > 100:
		log("WARN  Kijiji allows max 100 pages")
		args.pages = 100

	log("INFO  Scraping " + str(args.pages) + " pages")

	run(args.source)
